#! /usr/bin/env python


import os
import sys
import pandas as pd
import datetime
import pandas_datareader.data as web
from pandas_datareader.data import Options
from yahoo_finance import Share
from fredapi import Fred
from openpyxl import load_workbook
from optparse import OptionParser


pretty = False

parser = OptionParser()
parser.add_option("-v", "--file", action="store_true", dest='pretty', default=False)
(opts, args) = parser.parse_args()


#######################
# EQUITIES INPUTS

wb_filename = 'pandastry.xlsx'

sIndex_large =[{'SP500': 'fred'}, {'DJIA': 'fred'}, {'NASDAQCOM':"fred"},{ 'NASDAQ100':'fred'}]
sIndex_mid=  ['RU2000PR', 'RUMIDCAPTR']
sIndex_small = ['WILLSMLCAP']
sIndex_large_offset = 4
sIndex_mid_offset = 11
sIndex_small_offset = 19


volatility = ['WIV', 'STLFSI']
manufacturing = ['MNFCTRIRSA', 'MNFCTRSMSA', 'MNFCTRIMSA']
housing = ['ASPUS', 'MSPNHSUS', 'MSACSR', 'MNMFS', 'HSTFCM', 'HSTFC', 'HSTFFHAI', 'HSTFVAG', 'CSUSHPINSA']

fredKey = 'eb701cc93a5ec9bc22987c681c138b12'



########################


def print_x(msg=''):
	if(msg=='') | opts.pretty == False:
		return
	else:
		print(msg)


def dict_to_datapoints(lists=[]):
	dpoints=[]
	for d in lists:
		for key in d:
			vl = key
			sc = d[key]
			dp = datapoint(vl, sc)
			dpoints.append(dp)

	return(dpoints)



#########################



class wbook(object):

	def __init__(self):
		pass


class sheet(object):

	def __init__(self):
		pass


class component(object):

	def __init__(self, dp=[]):

		self.dpoints = dp
		print(len(self.dpoints))



		pass


class datapoint(object):

	def __init__(self, symbol, source):
		self.sym = symbol
		self.source = source
		if(self.source == 'fred'):
			self.fred = Fred(fredKey)
			self.fred_info(symbol)

	def fred_info(self, symbol):
		info = self.fred.get_series_info(symbol)
		self.title = info.title
		self.frequency = info.frequency_short.lower()
		self.units = info.units_short.lower()






#########################


class bot(object):
	

	def __init__(self):
		self.fred = Fred(api_key=fredKey)

		print("DailyUpdate\nv1.0\n...")

		print_x("\nStock indicies:\nLarge-cap: ")
		print_x(equity_large)

		print_x("Stock indicies:\nMid-cap: ")
		print_x(equity_mid)

		print_x("Stock indicies:\nSmall-cap:")
		print_x(equity_small)

		print_x("\n\nFilenames:")
		print_x("Stock indicies: " + equity_filename)

	def fredDownload(self, keyList=[]):

		#init dictionary array to hold series
		df = {}
		#create numPy series for each key in the given key list, add to array df
		for key in keyList:
			df[key] = self.fred.get_series(key, observation_start='2017-01-01', observation_end='2017-06-27')
		#create a data frame from array df
		df = pd.DataFrame(df)
		#sort descending by date
		sorted = df.sort_index(axis=0, ascending=False)
		print(sorted)
		return(sorted)


	def addTitles(self, keyList=[]):
		dicts = {}


	def dataReader(self, service, tickerlist=[]):
		start = datetime.datetime(2017,1,1)
		end = datetime.datetime(2017,1,27)
		df = web.DataReader(tickerlist, service, start, end)
		return(df)


	def largeCap(self):
		print_x("Downloading data for LARGE CAP STOCK INDICIES.....")
		df = self.dataReader('fred', equity_large)
		print_x("...SUCCESS")
		print_x("Downloading data for LARGE CAP STOCK INDICIES.....")
		print_x('...')
		print_x('Writing to excel.....')
		self.excel_write_exist(equity_filename,df, 22, large_offset)
		print_x("...SUCCESS")
		return(df)

	def midCap(self):
		print_x("Downloading data for MID CAP STOCK INDICIES......")
		df = self.dataReader('fred', equity_mid)
		print_x("...SUCCESS")
		print_x('Writing to excel.....')
		self.excel_write_exist(equity_filename,df, 22, mid_offset)
		print_x("...SUCCESS")
		return(df)

	def smallCap(self):
		print_x("Downloading data for SMALL CAP STOCK INDICIES......")
		df = self.dataReader('fred', equity_small)
		print_x("...SUCCESS")
		print_x('Writing to excel.....')
		self.excel_write_exist(equity_filename,df, 22, small_offset)
		print_x("...SUCCESS")
		return(df)

	def runSheets(self):
		pass




	def runComponent(self, datapoints=[]):


		pass







	def indicies(self):
		input("Press ENTER to update indicies sheet")
		print("\n\n\nDownloading data from indicies in indicies_list and extracting to excel")
		self.largeCap()
		self.midCap()
		self.smallCap()

	def excel_write_exist(self, filename, dFrame, row, col):

		book = load_workbook(filename)
		writer = pd.ExcelWriter(filename, engine='openpyxl')
		writer.book = book
		writer.sheets = dict((ws.title,ws) for ws in book.worksheets)
		df = dFrame
		df.to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=row, startcol=col)
		writer.save()
		return

	def excel_write_new(self, dFrame, sheetName,row, col):
		df = dFrame
		writer = pd.ExcelWriter('pandas_simple.xlsx', engine= 'xlsxwriter')
		df.to_excel(writer, sheet_name=sheetName, startrow=row, startcol=col)
		writer.save()



def main():
	#bot1= bot()

	dps = dict_to_datapoints(sIndex_large)
	comp = component(dps)

	#nm = 'SP500'
	#dp = datapoint('SP500', 'fred')
	#print(dp.title)

	#df = bot1.largeCap()
	#filename='pandastry.xlsx'
	#bot1.excel_write_exist(filename,df,22,4)
	#input('\n\n\nPress enter to begin...')
	#bot1.indicies()
	#print("\n\n\noperation complete.")


'''
	amzn = web.get_quote_yahoo('AMZN')
	print(amzn['last'])
	
	print(gdp)
	book = load_workbook('pandastry.xlsx')
	writer = pd.ExcelWriter('pandastry.xlsx', engine='openpyxl')
	writer.book = book
	writer.sheets = dict((ws.title,ws) for ws in book.worksheets)

	gdp.to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=22, startcol=4)

	writer.save()

	
	aapl = Options('aapl', 'yahoo')
	data = aapl.get_all_data()
	print(data.iloc[0:5, 0:5])
	'''
if __name__ == "__main__":
	main()


